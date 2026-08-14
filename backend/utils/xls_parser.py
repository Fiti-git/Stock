"""
XLS Import Engine for Arunalu Super Mart POS stock balance files.

Fixed column positions (0-indexed):
  Col 0  — item_code     (e.g. AR00003244, WGHE000380)
  Col 2  — item_name     (no header, position is fixed)
  Col 6  — cost_price    (LKR)
  Col 7  — selling_price (LKR)
  Col 8  — pos_quantity  (SIH — stock in hand)
"""

import re
from dataclasses import dataclass, field
from datetime import date
from typing import Optional

import openpyxl
import xlrd


# ---------------------------------------------------------------------------
# Column indices
#
# At least three POS export layouts exist in the wild and users can't be
# told which version of the report to run, so the parser must auto-detect:
#
#   "v1" (11 cols, includes a 'TTL LTR' column, alpha-prefix item codes):
#       0=item_code  2=name  5=TTL_LTR  6=COST  7=SELLING  8=SIH  9/10=values
#
#   "v2" (10 cols, 'TTL LTR' removed, alpha-prefix item codes):
#       0=item_code  2=name  4=TTL_LTR  5=COST  6=SELLING  7=SIH  8/9=values
#
#   "v3" (11 cols, includes 'EXP. DATE', digit-only item codes,
#         header LIES about column positions — header says cost=6 but the
#         actual numeric block is at cols 5-7 with col 8 blank):
#       0=item_code  1=name  5=COST  6=SELLING  7=SIH  9/10=values
#
# Detection strategy: find the first row whose col-0 matches an item-code
# pattern, then locate the actual SIH column by walking left from the
# rightmost numeric cells looking for three contiguous numbers. This
# overrides the header when they disagree (v3 case).
# ---------------------------------------------------------------------------
COL_ITEM_CODE = 0
COL_ITEM_NAME_DEFAULT = 2

LAYOUT_V1 = {"cost": 6, "selling": 7, "sih": 8, "name": 2}
LAYOUT_V2 = {"cost": 5, "selling": 6, "sih": 7, "name": 2}
LAYOUT_V3 = {"cost": 5, "selling": 6, "sih": 7, "name": 1}
DEFAULT_LAYOUT = LAYOUT_V1

# Accepts alpha-prefixed codes (AR00003244, WGHE000380, ASFAKFITI001) and
# pure-digit codes (013252, 009766). Requires at least three trailing
# digits to avoid false matches against tokens like "NO" or "ITEM".
ITEM_CODE_PATTERN = re.compile(r"^[A-Z]*\d{3,}$")


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------
@dataclass
class ParsedRow:
    item_code: str
    item_name: str
    cost_price: Optional[float]
    selling_price: Optional[float]
    pos_quantity: float
    category: str = ""


@dataclass
class ParseResult:
    rows: list = field(default_factory=list)
    snapshot_date: Optional[date] = None
    outlet_name: Optional[str] = None
    errors: list = field(default_factory=list)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _to_float(val) -> Optional[float]:
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _is_item_code(val: str) -> bool:
    return bool(val and ITEM_CODE_PATTERN.match(str(val).strip()))


def _normalise_cell(val) -> str:
    """Return cell value as a stripped string."""
    if val is None:
        return ""
    return str(val).strip()


# ---------------------------------------------------------------------------
# Row-type classification
# ---------------------------------------------------------------------------
def classify_row(cells: list, layout: dict = DEFAULT_LAYOUT) -> str:
    """Return one of: 'data', 'category', 'total', 'blank'."""
    col0 = _normalise_cell(cells[0] if len(cells) > 0 else "")
    sih_idx = layout["sih"]
    sih_cell = cells[sih_idx] if len(cells) > sih_idx else None

    if not any(_normalise_cell(c) for c in cells):
        return "blank"

    # 'TOTAL FOR :' label can appear in different columns depending on layout
    # (col 4 in v1, col 7 in v2). Scan all cells.
    for c in cells:
        if _normalise_cell(c).startswith("TOTAL FOR"):
            return "total"

    if _is_item_code(col0) and sih_cell is not None and sih_cell != "":
        try:
            float(sih_cell)
            return "data"
        except (TypeError, ValueError):
            pass

    if col0 and col0 == col0.upper() and not _is_item_code(col0):
        return "category"

    return "blank"


def _is_number(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _layout_from_header(rows: list):
    """
    Return a layout dict from the ITEM header row, or None if no usable
    header found. The header is ADVISORY only — caller validates against
    a sample data row and overrides if they disagree (v3 ships a header
    that doesn't match its own data layout).
    """
    for row in rows[:25]:
        normalised = [_normalise_cell(c).upper() for c in row]
        if "COST" in normalised and "SELLING" in normalised:
            cost_idx = normalised.index("COST")
            selling_idx = normalised.index("SELLING")
            sih_idx = normalised.index("SIH") if "SIH" in normalised else selling_idx + 1
            return {"cost": cost_idx, "selling": selling_idx, "sih": sih_idx, "name": COL_ITEM_NAME_DEFAULT}
    return None


def _layout_from_data(rows: list):
    """
    Probe the first item-code row to find the real numeric layout.

    POS exports always end every data row with two value columns
    (cost_value, selling_value). SIH sits one or two columns to the LEFT
    of cost_value (one in v1/v2, two in v3 because v3 leaves a blank
    cell between SIH and cost_value). SELLING and COST are immediately
    left of SIH.

    Returns a layout dict or None if no valid data row found.
    """
    for raw in rows:
        row = list(raw)
        if not row:
            continue
        c0 = _normalise_cell(row[0])
        if not _is_item_code(c0):
            continue

        # Find numeric column indices.
        num_indices = [i for i, v in enumerate(row) if _is_number(v)]
        if len(num_indices) < 3:
            continue

        # Rightmost two numerics = (cost_value, selling_value).
        cost_value_idx = num_indices[-2]

        # SIH is the rightmost numeric strictly left of cost_value such that
        # the two cells immediately left of it are also numeric (cost+selling).
        for sih_try in range(cost_value_idx - 1, max(-1, cost_value_idx - 4), -1):
            if sih_try < 2:
                break
            if _is_number(row[sih_try]) and _is_number(row[sih_try - 1]) and _is_number(row[sih_try - 2]):
                # Item name = first non-empty, non-numeric cell after col 0.
                name_idx = COL_ITEM_NAME_DEFAULT
                for i in range(1, min(len(row), 6)):
                    v = row[i]
                    if isinstance(v, str) and v.strip() and not _is_number(v):
                        name_idx = i
                        break
                return {
                    "cost": sih_try - 2,
                    "selling": sih_try - 1,
                    "sih": sih_try,
                    "name": name_idx,
                }
        # Found an item-code row but couldn't locate a 3-numeric run — give
        # up rather than guessing; the next item-code row may parse cleaner.
    return None


def _detect_layout(rows: list) -> dict:
    """
    Pick the column layout for this file. The data probe is the source of
    truth because it actually inspects values; the header is used only if
    the probe finds no item-code rows (e.g. an empty export).
    """
    data_layout = _layout_from_data(rows)
    if data_layout is not None:
        return data_layout
    header_layout = _layout_from_header(rows)
    if header_layout is not None:
        return header_layout
    return DEFAULT_LAYOUT


# ---------------------------------------------------------------------------
# File readers — normalise to list-of-lists regardless of format
# ---------------------------------------------------------------------------
def _read_xlsx(file) -> list:
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    wb.close()
    return rows


def _read_xls(file) -> list:
    data = file.read() if hasattr(file, "read") else open(file, "rb").read()
    try:
        wb = xlrd.open_workbook(file_contents=data)
    except Exception as e:
        raise ValueError(
            "This .xls file couldn't be opened. "
            "Please open it in Microsoft Excel, choose File → Save As, "
            "select \"Excel Workbook (.xlsx)\" as the format, save, and "
            "upload the new .xlsx file instead. "
            f"(Reason: {type(e).__name__})"
        ) from e
    ws = wb.sheet_by_index(0)
    rows = []
    for r in range(ws.nrows):
        row = []
        for c in range(ws.ncols):
            cell = ws.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_EMPTY:
                row.append(None)
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                row.append(cell.value)
            elif cell.ctype == xlrd.XL_CELL_DATE:
                row.append(xlrd.xldate_as_datetime(cell.value, wb.datemode).date())
            else:
                row.append(str(cell.value).strip())
        rows.append(row)
    return rows


def _get_all_rows(file, filename: str) -> list:
    name = filename.lower()
    if name.endswith(".xlsx"):
        return _read_xlsx(file)
    elif name.endswith(".xls"):
        if hasattr(file, "seek"):
            file.seek(0)
        return _read_xls(file)
    raise ValueError("Unsupported file format. Only .xls and .xlsx are accepted.")


# ---------------------------------------------------------------------------
# Header extraction
# ---------------------------------------------------------------------------
_HEADER_KEYWORDS = {
    "STOCK", "LOCATION", "DATE", "SELECTION", "ITEM", "REPORT",
    "ARUNALU", "SUPER", "MART", "NO", "AS", "AT", "BALANCE",
}


def _extract_headers(rows: list) -> tuple:
    """
    Scan first 20 rows for:
      "Date As At"   → snapshot_date
      "SUPER MARKET:" → outlet_name (legacy format)
      standalone uppercase location name (e.g. "AMPITIYA") → outlet_name (current format)
    Returns (snapshot_date: date | None, outlet_name: str | None)
    """
    from datetime import datetime

    snapshot_date = None
    outlet_name = None

    for row in rows[:20]:
        for i, cell in enumerate(row):
            text = _normalise_cell(cell)
            if "Date As At" in text or "DATE AS AT" in text:
                # Date may be several columns over — scan remaining cells for first non-empty
                for d in row[i + 1:]:
                    if d is None or _normalise_cell(d) == "":
                        continue
                    if isinstance(d, date):
                        snapshot_date = d
                    else:
                        try:
                            snapshot_date = datetime.strptime(
                                _normalise_cell(d), "%d/%m/%Y"
                            ).date()
                        except (ValueError, TypeError):
                            pass
                    break  # stop after first non-empty cell regardless

            if "SUPER MARKET:" in text:
                # Legacy format: value follows colon in same cell or next cell
                if ":" in text:
                    parts = text.split(":", 1)
                    outlet_name = parts[1].strip() if len(parts) > 1 else None
                elif i + 1 < len(row):
                    outlet_name = _normalise_cell(row[i + 1])

    # Current format: outlet name is a standalone uppercase cell in col 0
    # e.g. "AMPITIYA" appearing after the header block with no other content
    if outlet_name is None:
        for row in rows[:20]:
            col0 = _normalise_cell(row[0] if row else "")
            if not col0:
                continue
            # Must be all letters/spaces, uppercase, length > 2, no colon or digits
            if (
                col0 == col0.upper()
                and col0.replace(" ", "").isalpha()
                and len(col0) > 2
                and col0 not in _HEADER_KEYWORDS
                and not any(kw in col0.split() for kw in _HEADER_KEYWORDS)
            ):
                outlet_name = col0
                break

    return snapshot_date, outlet_name


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
def parse_xls(file, filename: str) -> ParseResult:
    """
    Parse a POS stock balance XLS/XLSX file.
    Returns a ParseResult with rows of ParsedRow and metadata.
    Does NOT touch the database.
    """
    result = ParseResult()

    try:
        all_rows = _get_all_rows(file, filename)
    except ValueError as e:
        result.errors.append(str(e))
        return result

    result.snapshot_date, result.outlet_name = _extract_headers(all_rows)
    layout = _detect_layout(all_rows)

    current_category = ""
    for raw_row in all_rows:
        # Pad row to at least 11 columns
        row = list(raw_row) + [None] * max(0, 11 - len(raw_row))
        kind = classify_row(row, layout)

        if kind == "category":
            current_category = _normalise_cell(row[COL_ITEM_CODE]).upper()
        elif kind == "data":
            item_code = _normalise_cell(row[COL_ITEM_CODE])
            name_idx = layout.get("name", COL_ITEM_NAME_DEFAULT)
            item_name = _normalise_cell(row[name_idx]) if name_idx < len(row) else ""
            pos_qty = _to_float(row[layout["sih"]])
            if pos_qty is None:
                continue
            result.rows.append(
                ParsedRow(
                    item_code=item_code,
                    item_name=item_name,
                    cost_price=_to_float(row[layout["cost"]]),
                    selling_price=_to_float(row[layout["selling"]]),
                    pos_quantity=pos_qty,
                    category=current_category,
                )
            )

    return result


def validate_file(file, filename: str) -> dict:
    """
    Run pre-import validation checks.
    Returns {"valid": bool, "errors": [...], "warnings": [...], "preview": {...}}

    Outlet is determined by user selection before upload — no header matching.
    Note: date is read from the file itself — no "must match today" check.
    The caller decides whether the date requires admin approval.
    """
    errors = []
    warnings = []

    result = parse_xls(file, filename)

    if result.errors:
        errors.extend(result.errors)
        return {"valid": False, "errors": errors, "warnings": warnings, "preview": {}}

    if result.snapshot_date is None:
        errors.append("Could not find 'Date As At' header in file.")

    if not result.rows:
        errors.append("No valid data rows found in the file.")

    preview = {
        "total_rows": len(result.rows),
        "snapshot_date": str(result.snapshot_date) if result.snapshot_date else None,
        "outlet_name": result.outlet_name,
    }

    return {
        "valid": len(errors) == 0,
        "errors": errors,
        "warnings": warnings,
        "preview": preview,
        "outlet_mismatch": None,
        "_parsed": result,
    }
