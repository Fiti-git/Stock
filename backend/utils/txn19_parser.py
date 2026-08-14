"""
Parser for the 19-column transaction reports shared by two types:

  * Direct Goods Received Note (GRN)
  * Returns to Supplier (RTS)

Banner layout is identical to the 17-col reports (Date From/To on r7,
outlet on r11). Detail grid starts on r15+ after the header on r13.

Column indices in the data rows (column labels and data line up cleanly
here — no shift like the Damage export):
  0  DoNo          "19,196.00"         document number
  1  Scode         "HINI0411"          supplier code
  2  CODE          "GCD0002790"        item code
  3  DATE          "2026-03-01 12:..." full datetime of the transaction
  4  Time          duplicate datetime — ignored
  5  INV NO        "024431"            supplier invoice no
  6  DESCRIPTION   "SPRITE PET 1.5"
  7  PckSIZE
  8  PRICE (cost)
  9  SPRICE (sell)
  10 PACKS
  11 UNITS (qty)   — actual delivered/returned quantity
  12 FREE QTY
  13 DISC %
  14 AMOUNT        = PRICE × UNITS
  15 USER          "SUSU"
  16 Tax %
  17 Tax Amount
  18 Tax Reg
"""

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Optional, Tuple

import openpyxl
import xlrd


# Column positions in data rows.
COL_DONO   = 0
COL_SCODE  = 1
COL_CODE   = 2
COL_DATE   = 3
COL_INV    = 5
COL_DESC   = 6
COL_PACK   = 7
COL_COST   = 8
COL_SELL   = 9
COL_PACKS  = 10
COL_QTY    = 11
COL_FREE   = 12
COL_DISC   = 13
COL_AMT    = 14
COL_USER   = 15
COL_TAX_P  = 16
COL_TAX_A  = 17
COL_TAX_R  = 18


_BANNER_STOPWORDS = {
    "ARUNALU", "SUPER", "MART", "REPORT", "LISTING", "DIRECT", "GOODS",
    "RECEIVED", "RETURNS", "SUPPLIER", "DATE", "FROM", "TO", "LOCATION",
    "STATUS", "PRINTED",
}
ITEM_CODE_PATTERN = re.compile(r"^[A-Z]{1,3}\w*\d+\w*$")


@dataclass
class Txn19Row:
    do_no: str
    supplier_code: str
    item_code: str
    description: str
    pack_size: str
    txn_date: date
    txn_time: str
    invoice_no: str
    cost_price: Optional[float]
    selling_price: Optional[float]
    packs: float
    qty: float
    free_qty: float
    disc_pct: float
    amount: float
    user_name: str
    tax_pct: float
    tax_amount: float
    tax_reg: str


@dataclass
class Txn19ParseResult:
    rows: list = field(default_factory=list)
    date_from: Optional[date] = None
    date_to: Optional[date] = None
    outlet_name: Optional[str] = None
    errors: list = field(default_factory=list)


# --------------------------------------------------------------------------- #
# File readers                                                                #
# --------------------------------------------------------------------------- #
def _read_xlsx(file) -> list:
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _read_xls(file) -> list:
    data = file.read() if hasattr(file, "read") else open(file, "rb").read()
    try:
        wb = xlrd.open_workbook(file_contents=data)
    except Exception as e:
        raise ValueError(
            "This .xls file couldn't be opened. "
            "Please open it in Microsoft Excel, choose File → Save As, "
            "select \"Excel Workbook (.xlsx)\" as the format, save, and "
            "upload the new .xlsx file instead. "
            f"(Reason: {type(e).__name__})"
        ) from e
    ws = wb.sheet_by_index(0)
    out = []
    for r in range(ws.nrows):
        row = []
        for c in range(ws.ncols):
            cell = ws.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_EMPTY:
                row.append(None)
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                row.append(cell.value)
            elif cell.ctype == xlrd.XL_CELL_DATE:
                row.append(xlrd.xldate_as_datetime(cell.value, wb.datemode))
            else:
                row.append(str(cell.value).strip())
        out.append(row)
    return out


def _get_rows(file, filename: str) -> list:
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return _read_xlsx(file)
    if name.endswith(".xls"):
        if hasattr(file, "seek"):
            file.seek(0)
        return _read_xls(file)
    raise ValueError("Unsupported file format. Only .xls and .xlsx are accepted.")


# --------------------------------------------------------------------------- #
# Helpers                                                                     #
# --------------------------------------------------------------------------- #
def _s(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _to_float(v, default=0.0) -> float:
    if v is None or v == "":
        return default
    try:
        # Strip thousands separators ("19,196.00" → "19196.00").
        if isinstance(v, str):
            v = v.replace(",", "")
        return float(v)
    except (TypeError, ValueError):
        return default


def _to_float_opt(v):
    if v is None or v == "":
        return None
    try:
        if isinstance(v, str):
            v = v.replace(",", "")
        return float(v)
    except (TypeError, ValueError):
        return None


def _parse_ddmmyyyy(s: str) -> Optional[date]:
    s = _s(s)
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _cell_as_datetime(v) -> Optional[datetime]:
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    s = _s(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    d = _parse_ddmmyyyy(s)
    return datetime(d.year, d.month, d.day) if d else None


def _extract_banner(rows: list) -> Tuple[Optional[date], Optional[date], Optional[str]]:
    date_from = date_to = None
    outlet_name = None
    for row in rows[:20]:
        for i, cell in enumerate(row):
            text = _s(cell).upper()
            if not text:
                continue
            if text.startswith("DATE FROM"):
                for d in row[i + 1:]:
                    parsed = _parse_ddmmyyyy(_s(d))
                    if parsed:
                        date_from = parsed
                        break
                    if isinstance(d, date):
                        date_from = d
                        break
            if text.startswith("DATE TO"):
                for d in row[i + 1:]:
                    parsed = _parse_ddmmyyyy(_s(d))
                    if parsed:
                        date_to = parsed
                        break
                    if isinstance(d, date):
                        date_to = d
                        break
    for row in rows[:20]:
        col0 = _s(row[0] if row else "")
        if not col0:
            continue
        if ":" in col0 and "SUPER MARKET" in col0.upper():
            outlet_name = col0.split(":", 1)[1].strip()
            break
        if (
            col0 == col0.upper()
            and col0.replace(" ", "").isalpha()
            and len(col0) > 2
            and not any(w in _BANNER_STOPWORDS for w in col0.split())
        ):
            outlet_name = col0
            break
    return date_from, date_to, outlet_name


def _looks_like_data_row(row: list) -> bool:
    do = _s(row[COL_DONO] if len(row) > COL_DONO else "")
    if not do or not do[0].isdigit():
        return False
    code = _s(row[COL_CODE] if len(row) > COL_CODE else "")
    if not code or not ITEM_CODE_PATTERN.match(code):
        return False
    dt = _cell_as_datetime(row[COL_DATE] if len(row) > COL_DATE else None)
    if dt is None:
        return False
    return True


def _format_doc(val) -> str:
    s = _s(val).replace(",", "")
    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
    except (TypeError, ValueError):
        pass
    return s


# --------------------------------------------------------------------------- #
# Public API                                                                  #
# --------------------------------------------------------------------------- #
def parse_txn19_xls(file, filename: str) -> Txn19ParseResult:
    result = Txn19ParseResult()
    try:
        all_rows = _get_rows(file, filename)
    except ValueError as e:
        result.errors.append(str(e))
        return result

    result.date_from, result.date_to, result.outlet_name = _extract_banner(all_rows)

    for raw in all_rows:
        row = list(raw) + [None] * max(0, 19 - len(raw))
        if not _looks_like_data_row(row):
            continue
        dt = _cell_as_datetime(row[COL_DATE])
        txn_date = dt.date() if dt else None
        txn_time = dt.strftime("%H:%M:%S") if dt else ""

        pack_raw = _s(row[COL_PACK])
        try:
            pf = float(pack_raw.replace(",", ""))
            if pf.is_integer():
                pack_raw = str(int(pf))
        except (TypeError, ValueError):
            pass

        result.rows.append(
            Txn19Row(
                do_no=_format_doc(row[COL_DONO]),
                supplier_code=_s(row[COL_SCODE]).upper()[:40],
                item_code=_s(row[COL_CODE]).upper()[:40],
                description=_s(row[COL_DESC])[:255],
                pack_size=pack_raw[:20],
                txn_date=txn_date,
                txn_time=txn_time[:20],
                invoice_no=_s(row[COL_INV])[:40],
                cost_price=_to_float_opt(row[COL_COST]),
                selling_price=_to_float_opt(row[COL_SELL]),
                packs=_to_float(row[COL_PACKS]),
                qty=_to_float(row[COL_QTY]),
                free_qty=_to_float(row[COL_FREE]),
                disc_pct=_to_float(row[COL_DISC]),
                amount=_to_float(row[COL_AMT]),
                user_name=_s(row[COL_USER])[:80],
                tax_pct=_to_float(row[COL_TAX_P]),
                tax_amount=_to_float(row[COL_TAX_A]),
                tax_reg=_s(row[COL_TAX_R])[:40],
            )
        )
    return result


def validate_txn19_file(file, filename: str) -> dict:
    errors = []
    warnings = []
    parsed = parse_txn19_xls(file, filename)

    if parsed.errors:
        errors.extend(parsed.errors)
    if parsed.date_from is None or parsed.date_to is None:
        errors.append("Could not find 'Date From' / 'Date To' in banner.")
    if not parsed.rows:
        errors.append("No valid data rows found in the file.")

    preview = {
        "total_rows": len(parsed.rows),
        "date_from": str(parsed.date_from) if parsed.date_from else None,
        "date_to": str(parsed.date_to) if parsed.date_to else None,
        "outlet_name": parsed.outlet_name,
        "total_amount": round(sum(r.amount for r in parsed.rows), 2),
    }
    return {
        "valid": not errors,
        "errors": errors,
        "warnings": warnings,
        "preview": preview,
        "_parsed": parsed,
    }
