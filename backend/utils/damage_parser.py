"""
Parser for 'Damaged/Wastage Entry Listing' XLS exports from Arunalu POS.

Banner layout (rows 0..12):
  r0   ARUNALU SUPER MART
  r2   <address>
  r4   DAMAGED/WASTAGE ENTRY LISTING ...
  r7   Date From : : 01/03/2026  Date To : : 31/03/2026
  r8   Location From : : 012  Location To : : 012
  r11  <OUTLET NAME>                 (standalone uppercase)

Detail grid (header on r15, real column indices we read from):
  0  DOC            "510.00"          document no.
  1  DATE           "03/03/2026"
  3  CODE           "GDF0009270"      item code
  4  DESCRIPTION    "SPRATS - 1KG"
  6  PACK SIZE      "1"
  7  CPRICE         "1000"
  8  SPRICE
  9  PACKS
  10 UNITS
  11 FREE QTY
  12 DISC %
  13 AMOUNT
  14 USER           "AKUMUDUNI"
  15 TIME           "1900-01-02 10:56:35"  (ignored — only the date matters)
"""

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Optional, Tuple

import openpyxl
import xlrd


# Column indices in the data rows. The XLS header labels are merged/shifted
# so the real column positions (verified against the sample: cost×qty=amount)
# differ from the header text by one column from col8 onwards.
COL_DOC   = 0
COL_DATE  = 1
COL_CODE  = 3
COL_DESC  = 4
COL_PACK  = 6
COL_CPRC  = 7
COL_SPRC  = 9   # header says 'PACKS' but contains the selling price
COL_QTY   = 11  # header says 'FREE QTY' but contains the damage quantity
COL_AMT   = 14  # header says 'USER' but contains cost × qty
COL_USER  = 15  # header says 'TIME' but contains the cashier/user name
COL_TIME  = 16  # unlabeled trailing column with the timestamp


# Keywords that rule out a cell from being a standalone outlet name on row 11.
_BANNER_STOPWORDS = {
    "ARUNALU", "SUPER", "MART", "REPORT", "LISTING", "DAMAGED",
    "DATE", "FROM", "TO", "LOCATION", "STATUS", "PRINTED",
}

ITEM_CODE_PATTERN = re.compile(r"^[A-Z]{1,3}\w*\d+\w*$")


@dataclass
class DamageRow:
    doc_no: str
    txn_date: date
    item_code: str
    description: str = ""
    pack_size: str = ""
    cost_price: Optional[float] = None
    selling_price: Optional[float] = None
    qty: float = 0
    amount: float = 0
    user_name: str = ""
    txn_time: str = ""


@dataclass
class DamageParseResult:
    rows: list = field(default_factory=list)
    date_from: Optional[date] = None
    date_to: Optional[date] = None
    outlet_name: Optional[str] = None
    errors: list = field(default_factory=list)


# -----------------------------------------------------------------------------
# File readers — copied from xls_parser so each type stays self-contained.
# -----------------------------------------------------------------------------
def _read_xlsx(file) -> list:
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _read_xls(file) -> list:
    data = file.read() if hasattr(file, "read") else open(file, "rb").read()
    try:
        wb = xlrd.open_workbook(file_contents=data)
    except Exception as e:
        raise ValueError(
            "This .xls file couldn't be opened. "
            "Please open it in Microsoft Excel, choose File → Save As, "
            "select \"Excel Workbook (.xlsx)\" as the format, save, and "
            "upload the new .xlsx file instead. "
            f"(Reason: {type(e).__name__})"
        ) from e
    ws = wb.sheet_by_index(0)
    out = []
    for r in range(ws.nrows):
        row = []
        for c in range(ws.ncols):
            cell = ws.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_EMPTY:
                row.append(None)
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                row.append(cell.value)
            elif cell.ctype == xlrd.XL_CELL_DATE:
                row.append(xlrd.xldate_as_datetime(cell.value, wb.datemode).date())
            else:
                row.append(str(cell.value).strip())
        out.append(row)
    return out


def _get_rows(file, filename: str) -> list:
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return _read_xlsx(file)
    if name.endswith(".xls"):
        if hasattr(file, "seek"):
            file.seek(0)
        return _read_xls(file)
    raise ValueError("Unsupported file format. Only .xls and .xlsx are accepted.")


# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------
def _s(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _to_float(v, default=0.0) -> float:
    if v is None or v == "":
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _to_float_opt(v) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _parse_ddmmyyyy(s: str) -> Optional[date]:
    s = _s(s)
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _cell_as_date(v) -> Optional[date]:
    if isinstance(v, date):
        return v
    if isinstance(v, datetime):
        return v.date()
    return _parse_ddmmyyyy(_s(v))


def _extract_banner(rows: list) -> Tuple[Optional[date], Optional[date], Optional[str]]:
    """Scan rows 0..20 for Date From / Date To / outlet name."""
    date_from = date_to = None
    outlet_name = None

    for row in rows[:20]:
        for i, cell in enumerate(row):
            text = _s(cell).upper()
            if not text:
                continue
            if text.startswith("DATE FROM"):
                for d in row[i + 1:]:
                    parsed = _cell_as_date(d)
                    if parsed:
                        date_from = parsed
                        break
            if text.startswith("DATE TO"):
                for d in row[i + 1:]:
                    parsed = _cell_as_date(d)
                    if parsed:
                        date_to = parsed
                        break

    # Outlet is typically a standalone uppercase string in col 0 before the
    # header row. Accept "ARUNALU SUPER MARKET:HINDAGA" or just "HINDAGA".
    for row in rows[:20]:
        col0 = _s(row[0] if row else "")
        if not col0 or ":" not in col0 and any(w in _BANNER_STOPWORDS for w in col0.upper().split()):
            # "ARUNALU SUPER MART" / report titles skip
            continue
        if ":" in col0 and "SUPER MARKET" in col0.upper():
            outlet_name = col0.split(":", 1)[1].strip()
            break
        if col0 == col0.upper() and col0.replace(" ", "").isalpha() and len(col0) > 2:
            if not any(w in _BANNER_STOPWORDS for w in col0.split()):
                outlet_name = col0
                break

    return date_from, date_to, outlet_name


def _looks_like_data_row(row: list) -> bool:
    """
    Data rows: col0 has a DOC no (numeric or numeric-like) AND col1 is a date
    AND col3 has an item code. This lets us skip banner/header/total rows.
    """
    doc = _s(row[COL_DOC] if len(row) > COL_DOC else "")
    if not doc:
        return False
    # doc_no is usually a float like "510.00" or "510" — must start with digit
    if not doc[0].isdigit():
        return False
    dt = _cell_as_date(row[COL_DATE] if len(row) > COL_DATE else None)
    if dt is None:
        return False
    code = _s(row[COL_CODE] if len(row) > COL_CODE else "")
    if not code or not ITEM_CODE_PATTERN.match(code):
        return False
    return True


def _format_doc(val) -> str:
    s = _s(val)
    # "510.00" -> "510"; "510.5" left as-is
    if s.endswith(".0") or s.endswith(".00"):
        try:
            return str(int(float(s)))
        except (TypeError, ValueError):
            pass
    # Strip trailing .0 on any whole number floats coming from xlrd
    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
    except (TypeError, ValueError):
        pass
    return s


# -----------------------------------------------------------------------------
# Public API
# -----------------------------------------------------------------------------
def parse_damage_xls(file, filename: str) -> DamageParseResult:
    result = DamageParseResult()

    try:
        all_rows = _get_rows(file, filename)
    except ValueError as e:
        result.errors.append(str(e))
        return result

    result.date_from, result.date_to, result.outlet_name = _extract_banner(all_rows)

    for raw in all_rows:
        row = list(raw) + [None] * max(0, 17 - len(raw))
        if not _looks_like_data_row(row):
            continue
        dt = _cell_as_date(row[COL_DATE])
        time_cell = row[COL_TIME]
        if isinstance(time_cell, datetime):
            time_str = time_cell.strftime("%H:%M:%S")
        else:
            time_str = _s(time_cell)
            # "1900-01-02 10:56:35" → "10:56:35"
            if " " in time_str and time_str.startswith("1900-"):
                time_str = time_str.split(" ", 1)[1]
        pack_raw = _s(row[COL_PACK])
        try:
            f = float(pack_raw)
            if f.is_integer():
                pack_raw = str(int(f))
        except (TypeError, ValueError):
            pass
        result.rows.append(
            DamageRow(
                doc_no=_format_doc(row[COL_DOC]),
                txn_date=dt,
                item_code=_s(row[COL_CODE]).upper(),
                description=_s(row[COL_DESC])[:255],
                pack_size=pack_raw[:20],
                cost_price=_to_float_opt(row[COL_CPRC]),
                selling_price=_to_float_opt(row[COL_SPRC]),
                qty=_to_float(row[COL_QTY]),
                amount=_to_float(row[COL_AMT]),
                user_name=_s(row[COL_USER])[:80],
                txn_time=time_str[:20],
            )
        )

    return result


def validate_damage_file(file, filename: str) -> dict:
    """Pre-import checks; no DB writes."""
    errors = []
    warnings = []
    parsed = parse_damage_xls(file, filename)

    if parsed.errors:
        errors.extend(parsed.errors)
    if parsed.date_from is None or parsed.date_to is None:
        errors.append("Could not find 'Date From' / 'Date To' in banner.")
    if not parsed.rows:
        errors.append("No valid damage rows found in the file.")

    # If data rows exist, tighten the covered range to the actual min/max of
    # the data. Banner says Mar 1-31 but if only 3 rows on Mar 3 are present,
    # we still store the banner window — that's what the user uploaded for.
    preview = {
        "total_rows": len(parsed.rows),
        "date_from": str(parsed.date_from) if parsed.date_from else None,
        "date_to": str(parsed.date_to) if parsed.date_to else None,
        "outlet_name": parsed.outlet_name,
        "total_amount": round(sum(r.amount for r in parsed.rows), 2),
    }
    return {
        "valid": not errors,
        "errors": errors,
        "warnings": warnings,
        "preview": preview,
        "_parsed": parsed,
    }
