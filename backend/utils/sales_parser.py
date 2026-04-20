"""
Parser for the 'Bill Listing Details' Sales XLS exports.

Banner: Date From / Date To on r10 (different from the txn exports' r7).
Header on r16, data from r17+. Data rows are interleaved with "Invoice
Total" subtotal rows that must be skipped.

Column layout (verified against real sample; header labels are shifted one
column to the right from the data starting at col 4):

  0  Date          "2026-03-04"
  2  Invoice No.   "2441282"
  4  Item Code     "GPN0020046"
  5  Description   "MAGGI NOODLES 158G"
  8  Cost          "200.91"
  9  Unit Price    "220"
  10 Qty           "1"
  11 Discount      "0"
  12 Amount        "220"
  13 Cashier       "SHIWANTHI"
  14 Time          "1900-01-02 20:38:00"   (only hh:mm:ss is meaningful)
"""

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Optional, Tuple

import openpyxl
import xlrd


COL_DATE    = 0
COL_CUSCODE = 1
COL_INV     = 2
COL_CODE    = 4
COL_DESC    = 5
COL_COST    = 8
COL_UNIT    = 9
COL_QTY     = 10
COL_DISC    = 11
COL_AMT     = 12
COL_CASH    = 13
COL_TIME    = 14

_BANNER_STOPWORDS = {
    "ARUNALU", "SUPER", "MART", "BILL", "LISTING", "REPORT", "DATE",
    "FROM", "TO", "LOCATION", "STATUS", "PRINTED", "INVOICE",
}
ITEM_CODE_PATTERN = re.compile(r"^[A-Z]{1,3}\w*\d+\w*$")


@dataclass
class SalesRow:
    invoice_no: str
    txn_date: date
    txn_time: str
    item_code: str
    description: str
    cust_code: str
    cost_price: Optional[float]
    unit_price: Optional[float]
    qty: float
    discount: float
    amount: float
    cashier: str


@dataclass
class SalesParseResult:
    rows: list = field(default_factory=list)
    date_from: Optional[date] = None
    date_to: Optional[date] = None
    outlet_name: Optional[str] = None
    errors: list = field(default_factory=list)


def _read_xlsx(file) -> list:
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _read_xls(file) -> list:
    data = file.read() if hasattr(file, "read") else open(file, "rb").read()
    wb = xlrd.open_workbook(file_contents=data)
    ws = wb.sheet_by_index(0)
    out = []
    for r in range(ws.nrows):
        row = []
        for c in range(ws.ncols):
            cell = ws.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_EMPTY:
                row.append(None)
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                row.append(cell.value)
            elif cell.ctype == xlrd.XL_CELL_DATE:
                row.append(xlrd.xldate_as_datetime(cell.value, wb.datemode))
            else:
                row.append(str(cell.value).strip())
        out.append(row)
    return out


def _get_rows(file, filename: str) -> list:
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return _read_xlsx(file)
    if name.endswith(".xls"):
        if hasattr(file, "seek"):
            file.seek(0)
        return _read_xls(file)
    raise ValueError("Unsupported file format. Only .xls and .xlsx are accepted.")


def _s(v) -> str:
    return "" if v is None else str(v).strip()


def _to_float(v, default=0.0) -> float:
    if v is None or v == "":
        return default
    try:
        if isinstance(v, str):
            v = v.replace(",", "")
        return float(v)
    except (TypeError, ValueError):
        return default


def _to_float_opt(v):
    if v is None or v == "":
        return None
    try:
        if isinstance(v, str):
            v = v.replace(",", "")
        return float(v)
    except (TypeError, ValueError):
        return None


def _cell_as_date(v) -> Optional[date]:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _s(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _cell_as_time(v) -> str:
    if isinstance(v, datetime):
        return v.strftime("%H:%M:%S")
    s = _s(v)
    if " " in s and s.startswith("1900-"):
        return s.split(" ", 1)[1]
    if s.startswith("1900-"):
        return ""
    return s


def _extract_banner(rows: list) -> Tuple[Optional[date], Optional[date], Optional[str]]:
    date_from = date_to = None
    outlet_name = None
    for row in rows[:20]:
        for i, cell in enumerate(row):
            text = _s(cell).upper()
            if not text:
                continue
            if text.startswith("DATE FROM"):
                for d in row[i + 1:]:
                    parsed = _cell_as_date(d)
                    if parsed:
                        date_from = parsed
                        break
            if text.startswith("DATE TO"):
                for d in row[i + 1:]:
                    parsed = _cell_as_date(d)
                    if parsed:
                        date_to = parsed
                        break
    # Outlet — "Location : ARUNALU SUPER MARKET:HINDAGALA"
    for row in rows[:25]:
        for i, cell in enumerate(row):
            text = _s(cell)
            if text.upper().startswith("LOCATION :"):
                for nxt in row[i + 1:]:
                    txt = _s(nxt)
                    if txt and ":" in txt and "SUPER MARKET" in txt.upper():
                        outlet_name = txt.split(":", 1)[1].strip()
                        return date_from, date_to, outlet_name
                    if txt and txt == txt.upper() and txt.replace(" ", "").isalpha():
                        outlet_name = txt
                        return date_from, date_to, outlet_name
    # Fallback: find uppercase standalone token
    for row in rows[:25]:
        col1 = _s(row[1] if len(row) > 1 else "")
        if col1 and ":" in col1 and "SUPER MARKET" in col1.upper():
            outlet_name = col1.split(":", 1)[1].strip()
            return date_from, date_to, outlet_name
    return date_from, date_to, outlet_name


def _looks_like_data_row(row: list) -> bool:
    # Data rows have a real date in col 0, an invoice in col 2, an item code in col 4.
    d = _cell_as_date(row[COL_DATE] if len(row) > COL_DATE else None)
    if not d:
        return False
    inv = _s(row[COL_INV] if len(row) > COL_INV else "")
    if not inv or not inv[0].isdigit():
        return False
    code = _s(row[COL_CODE] if len(row) > COL_CODE else "")
    if not code or not ITEM_CODE_PATTERN.match(code):
        return False
    return True


def _format_invoice(val) -> str:
    s = _s(val)
    try:
        f = float(s.replace(",", ""))
        if f.is_integer():
            return str(int(f))
    except (TypeError, ValueError):
        pass
    return s


def parse_sales_xls(file, filename: str) -> SalesParseResult:
    result = SalesParseResult()
    try:
        all_rows = _get_rows(file, filename)
    except ValueError as e:
        result.errors.append(str(e))
        return result

    result.date_from, result.date_to, result.outlet_name = _extract_banner(all_rows)

    for raw in all_rows:
        row = list(raw) + [None] * max(0, 15 - len(raw))
        if not _looks_like_data_row(row):
            continue
        result.rows.append(
            SalesRow(
                invoice_no=_format_invoice(row[COL_INV])[:40],
                txn_date=_cell_as_date(row[COL_DATE]),
                txn_time=_cell_as_time(row[COL_TIME])[:20],
                item_code=_s(row[COL_CODE]).upper()[:40],
                description=_s(row[COL_DESC])[:255],
                cust_code=_s(row[COL_CUSCODE])[:40],
                cost_price=_to_float_opt(row[COL_COST]),
                unit_price=_to_float_opt(row[COL_UNIT]),
                qty=_to_float(row[COL_QTY]),
                discount=_to_float(row[COL_DISC]),
                amount=_to_float(row[COL_AMT]),
                cashier=_s(row[COL_CASH])[:80],
            )
        )
    return result


def validate_sales_file(file, filename: str) -> dict:
    errors, warnings = [], []
    parsed = parse_sales_xls(file, filename)
    if parsed.errors:
        errors.extend(parsed.errors)
    if parsed.date_from is None or parsed.date_to is None:
        errors.append("Could not find 'Date From' / 'Date To' in the banner.")
    if not parsed.rows:
        errors.append("No valid sales rows found in the file.")
    preview = {
        "total_rows": len(parsed.rows),
        "date_from": str(parsed.date_from) if parsed.date_from else None,
        "date_to": str(parsed.date_to) if parsed.date_to else None,
        "outlet_name": parsed.outlet_name,
        "total_amount": round(sum(r.amount for r in parsed.rows), 2),
    }
    return {
        "valid": not errors,
        "errors": errors,
        "warnings": warnings,
        "preview": preview,
        "_parsed": parsed,
    }
