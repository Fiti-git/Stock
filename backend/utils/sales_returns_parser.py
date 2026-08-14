"""
Parser for the 'Sales Returns Details (With Reason)' XLS export.

Banner: Date From/To on r8, Location on r11. Header on r17, data from r19+.
Unlike the Sales bill-listing export, column indices in the data rows line
up directly with the header labels (no shift).

  0  Date         "2026-03-01"
  2  Inv#         "1360755"
  3  PLU          "GDE0249919"          item code
  5  Barcode      "GDE0249919"
  6  Description
  7  Quantity                           typically negative
  8  Cost         "998"
  9  Gross Value                        typically negative (return value)
  10 Remarks                            can be blank
  11 user         "KASUN"
  12 DateTime     "2026-03-01 12:33:30"
"""

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Optional, Tuple

import openpyxl
import xlrd


COL_DATE   = 0
COL_INV    = 2
COL_PLU    = 3
COL_MEMBER = 4
COL_BAR    = 5
COL_DESC   = 6
COL_QTY    = 7
COL_COST   = 8
COL_GROSS  = 9
COL_REMARK = 10
COL_USER   = 11
COL_DT     = 12

ITEM_CODE_PATTERN = re.compile(r"^[A-Z]{1,3}\w*\d+\w*$")


@dataclass
class SalesReturnRow:
    invoice_no: str
    txn_date: date
    txn_time: str
    item_code: str
    barcode: str
    description: str
    member: str
    qty: float
    cost_price: Optional[float]
    gross_value: float
    remarks: str
    user_name: str


@dataclass
class SalesReturnParseResult:
    rows: list = field(default_factory=list)
    date_from: Optional[date] = None
    date_to: Optional[date] = None
    outlet_name: Optional[str] = None
    errors: list = field(default_factory=list)


def _read_xlsx(file) -> list:
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _read_xls(file) -> list:
    data = file.read() if hasattr(file, "read") else open(file, "rb").read()
    try:
        wb = xlrd.open_workbook(file_contents=data)
    except Exception as e:
        raise ValueError(
            "This .xls file couldn't be opened. "
            "Please open it in Microsoft Excel, choose File → Save As, "
            "select \"Excel Workbook (.xlsx)\" as the format, save, and "
            "upload the new .xlsx file instead. "
            f"(Reason: {type(e).__name__})"
        ) from e
    ws = wb.sheet_by_index(0)
    out = []
    for r in range(ws.nrows):
        row = []
        for c in range(ws.ncols):
            cell = ws.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_EMPTY:
                row.append(None)
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                row.append(cell.value)
            elif cell.ctype == xlrd.XL_CELL_DATE:
                row.append(xlrd.xldate_as_datetime(cell.value, wb.datemode))
            else:
                row.append(str(cell.value).strip())
        out.append(row)
    return out


def _get_rows(file, filename: str) -> list:
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return _read_xlsx(file)
    if name.endswith(".xls"):
        if hasattr(file, "seek"):
            file.seek(0)
        return _read_xls(file)
    raise ValueError("Unsupported file format. Only .xls and .xlsx are accepted.")


def _s(v) -> str:
    return "" if v is None else str(v).strip()


def _to_float(v, default=0.0) -> float:
    if v is None or v == "":
        return default
    try:
        if isinstance(v, str):
            v = v.replace(",", "")
        return float(v)
    except (TypeError, ValueError):
        return default


def _to_float_opt(v):
    if v is None or v == "":
        return None
    try:
        if isinstance(v, str):
            v = v.replace(",", "")
        return float(v)
    except (TypeError, ValueError):
        return None


def _cell_as_date(v) -> Optional[date]:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _s(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _cell_as_datetime(v) -> Optional[datetime]:
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    s = _s(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _extract_banner(rows: list) -> Tuple[Optional[date], Optional[date], Optional[str]]:
    date_from = date_to = None
    outlet_name = None
    for row in rows[:20]:
        for i, cell in enumerate(row):
            text = _s(cell).upper()
            if not text:
                continue
            if text.startswith("DATE FROM"):
                for d in row[i + 1:]:
                    parsed = _cell_as_date(d)
                    if parsed:
                        date_from = parsed
                        break
            if text.startswith("DATE TO"):
                for d in row[i + 1:]:
                    parsed = _cell_as_date(d)
                    if parsed:
                        date_to = parsed
                        break
    # outlet — the "location :" line lives around rows 11–18 and the outlet
    # name typically appears in col 1 after a location code in col 0.
    for row in rows[:25]:
        col1 = _s(row[1] if len(row) > 1 else "")
        if col1 and ":" in col1 and "SUPER MARKET" in col1.upper():
            outlet_name = col1.split(":", 1)[1].strip()
            break
    return date_from, date_to, outlet_name


def _looks_like_data_row(row: list) -> bool:
    d = _cell_as_date(row[COL_DATE] if len(row) > COL_DATE else None)
    if not d:
        return False
    inv = _s(row[COL_INV] if len(row) > COL_INV else "")
    if not inv or not inv[0].isdigit():
        return False
    code = _s(row[COL_PLU] if len(row) > COL_PLU else "")
    if not code or not ITEM_CODE_PATTERN.match(code):
        return False
    return True


def _format_invoice(val) -> str:
    s = _s(val)
    try:
        f = float(s.replace(",", ""))
        if f.is_integer():
            return str(int(f))
    except (TypeError, ValueError):
        pass
    return s


def parse_sales_returns_xls(file, filename: str) -> SalesReturnParseResult:
    result = SalesReturnParseResult()
    try:
        all_rows = _get_rows(file, filename)
    except ValueError as e:
        result.errors.append(str(e))
        return result

    result.date_from, result.date_to, result.outlet_name = _extract_banner(all_rows)

    for raw in all_rows:
        row = list(raw) + [None] * max(0, 13 - len(raw))
        if not _looks_like_data_row(row):
            continue
        dt_full = _cell_as_datetime(row[COL_DT])
        time_str = dt_full.strftime("%H:%M:%S") if dt_full else ""
        result.rows.append(
            SalesReturnRow(
                invoice_no=_format_invoice(row[COL_INV])[:40],
                txn_date=_cell_as_date(row[COL_DATE]),
                txn_time=time_str[:20],
                item_code=_s(row[COL_PLU]).upper()[:40],
                barcode=_s(row[COL_BAR])[:40],
                description=_s(row[COL_DESC])[:255],
                member=_s(row[COL_MEMBER])[:60],
                qty=_to_float(row[COL_QTY]),
                cost_price=_to_float_opt(row[COL_COST]),
                gross_value=_to_float(row[COL_GROSS]),
                remarks=_s(row[COL_REMARK])[:255],
                user_name=_s(row[COL_USER])[:80],
            )
        )
    return result


def validate_sales_returns_file(file, filename: str) -> dict:
    errors, warnings = [], []
    parsed = parse_sales_returns_xls(file, filename)
    if parsed.errors:
        errors.extend(parsed.errors)
    if parsed.date_from is None or parsed.date_to is None:
        errors.append("Could not find 'Date From' / 'Date To' in the banner.")
    if not parsed.rows:
        errors.append("No valid return rows found in the file.")
    preview = {
        "total_rows": len(parsed.rows),
        "date_from": str(parsed.date_from) if parsed.date_from else None,
        "date_to": str(parsed.date_to) if parsed.date_to else None,
        "outlet_name": parsed.outlet_name,
        "total_amount": round(sum(r.gross_value for r in parsed.rows), 2),
    }
    return {
        "valid": not errors,
        "errors": errors,
        "warnings": warnings,
        "preview": preview,
        "_parsed": parsed,
    }
